import tkinter as tk
from datetime import datetime
import os
from dotenv import load_dotenv
import openpyxl
import homePage


def create_directory(dir_path):
    if not os.path.exists(dir_path):
        os.mkdir(path=dir_path)


def add_task(task, task_list):
    task_list.insert(tk.END, task.get())


def delete_task(frame, task_list, title, data_dir):
    todo_title = title.get(1.0, tk.END).strip()

    try:
        selected_index = task_list.curselection()[0]
        task_list.delete(selected_index)
        todo_task_list = task_list.get(0, tk.END)

        if os.path.exists(f"{data_dir}\\{todo_title}.xlsx"):
            todo_excel = openpyxl.load_workbook(f"{data_dir}\\{todo_title}.xlsx")
            sheet = todo_excel[todo_title]

            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                try:
                    todo_task_list.index(row[0])
                except ValueError:
                    sheet.delete_rows(row_index+1)

            todo_excel.save(f"{data_dir}\\{todo_title}.xlsx")

    except IndexError:
        popup = tk.Toplevel(frame)
        popup.title("No Task Selected")
        popup.geometry("500x100")

        # Add a label to the popup window
        label = tk.Label(popup, text="Please select a task in the task list before attempting the delete operation!")
        label.pack(pady=10)

        # Add a button to close the popup
        close_button = tk.Button(popup, text="Close", command=popup.destroy)
        close_button.pack(pady=5)


def save_todo(data_dir, title, task_list):
    todo_title = title.get(1.0, tk.END).strip()
    todo_task_list = task_list.get(0, tk.END)

    if not os.path.exists(f"{data_dir}\\{todo_title}.xlsx"):
        todo_excel = openpyxl.Workbook()
        todo_excel.remove(todo_excel["Sheet"])
        sheet = todo_excel.create_sheet(todo_title)

        for task in todo_task_list:
            sheet.append([task])
    else:
        todo_excel = openpyxl.load_workbook(f"{data_dir}\\{todo_title}.xlsx")
        sheet = todo_excel[todo_title]

        sheet_rows = [rows[0] for rows in sheet.iter_rows(values_only=True)]

        for task in todo_task_list:
            try:
                sheet_rows.index(task)
            except ValueError:
                sheet.append([task])

        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            try:
                todo_task_list.index(row[0])
            except ValueError:
                task_list.insert(tk.END, row[0])

    todo_excel.save(f"{data_dir}\\{todo_title}.xlsx")


def create_new_todo(todo_title="", task_list=None):

    load_dotenv()

    # Create the main window
    root = tk.Tk()
    root.title("Profess")
    root.geometry("600x700")  # Set the window size

    # Create a frame widget
    frame = tk.Frame(root, bg="lightblue", bd=5, relief=tk.FLAT)
    frame.pack(fill=tk.BOTH, expand=True)  # Pack the frame with padding

    today = datetime.today()
    today_title_format = today.strftime('%d %B %Y, %I:%M %p')

    title_lbl = tk.Label(frame, text=f"Date: {today_title_format}",
                         bg="lightblue", fg="royal blue",
                         font=("Harlow Solid Italic", 20))
    title_lbl.grid(row=0, column=0, sticky="nsew", columnspan=3)

    todo_title_label = tk.Label(frame, text=f"Title",
                                bg="lightblue", fg="royal blue",
                                font=("Comic Sans MS", 14, "bold"))
    todo_title_label.grid(row=1, column=0, padx=20, sticky="w", columnspan=3)
    todo_title_text = tk.Text(frame, height=2)
    todo_title_text.insert(tk.END, todo_title)

    todo_title_text.grid(row=2, column=0, padx=20, sticky="ew", columnspan=3)

    todo_text_label = tk.Label(frame, text=f"Create a To Do List...",
                               bg="lightblue", fg="royal blue",
                               font=("Comic Sans MS", 14, "bold"))
    todo_text_label.grid(row=4, column=0, sticky="w", padx=20, columnspan=3)

    task_entry = tk.Entry(frame, width=25)
    task_entry.insert(0, "New Task")
    task_entry.grid(row=5, column=0, padx=(20, 0), sticky="ew")

    todo_list = tk.Listbox(frame, height=6)
    todo_list.grid(row=6, column=0, padx=(20, 20), sticky="ew", columnspan=3)
    if task_list is not None:
        for task in task_list:
            todo_list.insert(tk.END, task)

    today_directory_format = today.strftime('%Y%m%d')
    journal_data_dir = os.getenv("JOURNAL_DATA")
    journal_data_today_dir = f"{journal_data_dir}To Do\\{today_directory_format}"

    add_task_button = tk.Button(frame,
                                text="Add Task",
                                font=("Comic Sans MS", 10),
                                command=lambda: [add_task(task_entry, todo_list)])
    add_task_button.grid(row=5, column=1, padx=(0, 20), sticky="e")

    delete_task_button = tk.Button(frame,
                                   text="Delete Task",
                                   font=("Comic Sans MS", 10),
                                   command=lambda: [delete_task(frame, todo_list, todo_title_text, journal_data_today_dir)])
    delete_task_button.grid(row=5, column=2, sticky="w")

    save_button = tk.Button(frame,
                            text="Save",
                            font=("Comic Sans MS", 12),
                            command=lambda: [create_directory(f"{journal_data_dir}To Do"),
                                             create_directory(journal_data_today_dir),
                                             save_todo(journal_data_today_dir,
                                                       todo_title_text,
                                                       todo_list)])
    save_button.grid(row=7, column=0, padx=(40, 40), sticky="ew")

    save_and_close_button = tk.Button(frame,
                                      text="Save & Close",
                                      font=("Comic Sans MS", 12),
                                      command=lambda: [create_directory(f"{journal_data_dir}Entries"),
                                                       create_directory(journal_data_today_dir),
                                                       save_todo(journal_data_today_dir,
                                                                 todo_title_text,
                                                                 todo_list),
                                                       root.destroy(),
                                                       homePage.launch_home_page()])
    save_and_close_button.grid(row=7, column=1, padx=(40, 40), sticky="ew")

    exit_button = tk.Button(frame,
                            text="Exit",
                            font=("Comic Sans MS", 12),
                            command=lambda: [root.destroy(), homePage.launch_home_page()])
    exit_button.grid(row=7, column=2, padx=(40, 40), sticky="ew")

    frame.rowconfigure(0, weight=1)
    frame.rowconfigure(1, weight=1)
    frame.rowconfigure(2, weight=1)
    frame.rowconfigure(3, weight=1)
    frame.rowconfigure(4, weight=1)
    frame.rowconfigure(5, weight=1)
    frame.rowconfigure(6, weight=1)
    frame.rowconfigure(7, weight=1)
    frame.columnconfigure(0, weight=1)
    frame.columnconfigure(1, weight=1)
    frame.columnconfigure(2, weight=1)

    # Start the main event loop
    root.mainloop()


if __name__ == '__main__':
    create_new_todo()
